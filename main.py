import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from datetime import datetime
from scraper import scrape_linkedin_jobs
from analyzer import analyze_all_jobs, get_cost_summary
import subprocess
import json
import os


def save_to_excel(jobs: list, cost: dict):
    wb = openpyxl.Workbook()
    ws: Worksheet = wb.active  # type: ignore
    ws.title = "Jobs"

    headers = ["Job Title", "Company", "Location", "Match %", "Reason", "Link"]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    top_jobs = jobs[:50]
    for row, job in enumerate(top_jobs, 2):
        match = job.get("match_percentage", 0)

        if match >= 80:
            fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        elif match >= 50:
            fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        else:
            fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        values = [
            job.get("title", "N/A"),
            job.get("company", "N/A"),
            job.get("location", "N/A"),
            f"{match}%",
            job.get("reason", ""),
            job.get("link", "N/A"),
        ]

        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.fill = fill
            cell.alignment = Alignment(horizontal="left")

    cost_row = len(top_jobs) + 3
    ws.cell(row=cost_row, column=1, value="Run Cost Summary").font = Font(bold=True)
    ws.cell(row=cost_row + 1, column=1, value="Input Tokens")
    ws.cell(row=cost_row + 1, column=2, value=cost["input_tokens"])
    ws.cell(row=cost_row + 2, column=1, value="Output Tokens")
    ws.cell(row=cost_row + 2, column=2, value=cost["output_tokens"])
    ws.cell(row=cost_row + 3, column=1, value="Total Cost (USD)")
    ws.cell(row=cost_row + 3, column=2, value=f"${cost['total_cost_usd']}")

    column_widths = [35, 25, 20, 12, 50, 50]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    filename = f"jobs_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    wb.save(filename)
    return filename


def save_to_json(jobs: list, cost: dict):
    data = {
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "total_jobs_found": len(jobs),
        "cost": cost,
        "jobs": jobs[:50]
    }

    # שמירה גם בתיקיית job-agent וגם ב-frontend/public
    base_dir = os.path.dirname(os.path.abspath(__file__))
    frontend_public = os.path.join(base_dir, "frontend", "public", "jobs.json")
    local_json = os.path.join(base_dir, "jobs.json")

    with open(local_json, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    if os.path.exists(os.path.dirname(frontend_public)):
        with open(frontend_public, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        print(f"JSON saved to frontend: {frontend_public}")

    print(f"JSON saved: {local_json}")
    return local_json


def main():
    print("Starting LinkedIn job scan...")

    jobs = scrape_linkedin_jobs(
        keywords="Full Stack Developer",
        location="Israel",
        num_pages=1
    )

    if not jobs:
        print("No jobs found. Try again later.")
        return

    print(f"\nAnalyzing {len(jobs)} jobs with Claude...\n")
    analyzed_jobs = analyze_all_jobs(jobs)
    cost = get_cost_summary()

    print("\nGenerating Excel file...")
    excel_file = save_to_excel(analyzed_jobs, cost)
    print(f"Excel saved: {excel_file}")

    print("\nGenerating JSON...")
    save_to_json(analyzed_jobs, cost)

    print("\n--- TOP 5 JOBS ---")
    for job in analyzed_jobs[:5]:
        print(f"  {job['match_percentage']}% | {job['title']} @ {job['company']}")
        print(f"         {job.get('location', '')} | {job.get('reason', '')}")
        print()

    print(f"--- COST ---")
    print(f"  Input tokens:  {cost['input_tokens']:,}")
    print(f"  Output tokens: {cost['output_tokens']:,}")
    print(f"  Total cost:    ${cost['total_cost_usd']}")

    subprocess.run(["open", excel_file])


if __name__ == "__main__":
    main()